import os
import pandas as pd
import glob
import numpy as np

import matplotlib.pyplot as plt
import openpyxl

'''
Preset Stuff
File Path
'''
pd.set_option('display.float_format', lambda x: '%9.4f' %x)
############# CHANGE ME ##############
files_path = (r"G:\Shared drives\TNEL - UMN\Project related material\PLASTICITY\Plasticity\DATA\OB51F")
'''
^Main Important Change, if CSV files from autoERP_temp are saved in right \csvs folder in file_path, everything should autogenerate
Only prereq is changing the filepath, and saving the right data IN THE RIGHT ORDER using autoERP_temp
IMPORTANT: MAKE SURE THE FIRST DATA POINTS FOR IL AND BLA ARE THE FIRST ONES RECORDED, THAT'S HOW THE PERCENTAGES ARE CALCULATED 
'''
RatSub = 'OB51F'
chanToSkip = [] # If any known bad channels to skip
######################################


csvs_path = (files_path + "\csvs")
read_files = sorted(glob.glob(os.path.join(csvs_path,"*.csv")))

# Changed array to be called dfs
dfs = []
paths = []
for files in read_files:
    OB_data = pd.read_csv(files, header =None, names = ['Index','Slope','Amplitude', 'AUC'])
    dfs.append(OB_data)
    paths.append(files)
    ###print(files)

# Create a channels array for each dataframe
channels = [0] * len(dfs)
for j, df in enumerate(dfs):  # Loop over each dataframe
    # Find the index of rows saying "channel"
    indexDf = df['Index'].dropna()
    # print(indexDf)
    chanIndexes = indexDf.index[indexDf.str.contains('Channel')].tolist()
    # Create channels array for this dataframe
    channels[j] = [0] * len(chanIndexes)
    for i, n in enumerate(chanIndexes):  # Loop over channel Indexes
        if i == len(chanIndexes) - 1:  # last channel, go to end of file
            # From index[i] to end (-1)
            channels[j][i] = df.iloc[chanIndexes[i] + 2:-1]
            # Change to numeric because it is only numbers now. Lets us do math.
            channels[j][i] = channels[j][i].apply(pd.to_numeric)
        else:  # All other channels. Go from channel index to the next channel index
            # From index[i]+2 (skip the channelX and header lines) to index[i+1] (next channel starting point)
            channels[j][i] = df.iloc[chanIndexes[i] + 2:chanIndexes[i + 1]]
            # Change to numeric because it is only numbers now. Lets us do math.
            channels[j][i] = channels[j][i].apply(pd.to_numeric)

# Get channel data by channels[index of dataframe][channel - 1]
# So channels[0][0] would be the data from the first file and channel 1.
# Heres how to get the mean
# print(channels[1][0]['Slope'].mean())


ilDf = []
blaDf = []
ilMean = []
blaMean = []
for j in range(len(dfs)):
    if 'BLA' in paths[j]:
        ilDf = channels[j][0]
        for i in range(0, 8):
            if channels[j][i]['Index'].count() > 35 and i not in chanToSkip:
                ilDf = ilDf.append(channels[j][i], ignore_index=True)
        # ilMean.append([ilDf[j]['Slope'].mean(), ilDf[j]['Amplitude'].mean(), ilDf[j]['AUC' ].mean()])
        ilMean.append(ilDf['AUC'].mean())
    else:
        blaDf = channels[j][8]
        for i in range(9, 16):
            if channels[j][i]['Index'].count() > 35 and i not in chanToSkip:
                blaDf = blaDf.append(channels[j][i], ignore_index=True)
        blaMean.append(blaDf['AUC'].mean())
###print('ilmean', ilMean)
###print('blamean', blaMean)

# Example code to iterate through all data and channels
# for data in range(len(dfs)):
# for chan in range(16):
# channels[data][chan]['AUC'].mean()

stimType = []
meanIt = 0
for i in range(0, len(paths), 2):
    if 'ERP_PRE' in paths[i]:
        if i != 0:
            if i < len(paths) - 2:
                ilMean.insert(meanIt, ilMean[meanIt])
                blaMean.insert(meanIt, blaMean[meanIt])
                meanIt += 1
            stimType.append('ERP_24')
        if i < len(paths) - 2:
            stimType.append('ERP_PRE')

    elif 'ERP_5' in paths[i]:
        stimType.append('ERP_5')
    elif 'ERP_30' in paths[i]:
        stimType.append('ERP_30')

    meanIt += 1

###print(stimType)
###print('--')
###print(ilMean)

IL = pd.DataFrame(ilMean, columns=['AUC - IL'])
# print(IL)
# IL1 = IL.T

BLA = pd.DataFrame(blaMean, columns=['AUC - BLA'])
# print('--')
# print(BLA)

joined = []
for i in range(len(ilMean)):
    joined.append([stimType[i], ilMean[i], blaMean[i]])

ILBLA = pd.DataFrame(joined, columns=['Stim Type', 'AUC-IL', 'AUC-BLA'])
###print(ILBLA)

ILBLA.to_csv(files_path + '\Results_' + RatSub + '.csv')



###################################################################################################################################Sufians
#this method doesn't work if the excel file is open since temp ~$file opens as well in FileStream
#if the error "df values not defined shows up", that's because no csv file exists in files path, check where it was saved (this shouldnt be a problem anymore though)

results = glob.glob(os.path.join(files_path,"*.csv")) # becomes saved as array, even though 1 file xlsx

ilPercent = []
blaPercent = []


for file in results: # thus making the for loop necessary
    data = pd.read_csv(file)#, header=None, names=['Stim Type', 'AUC-IL', 'AUC-BLA'])
    df = pd.DataFrame(data, columns=['Stim Type', 'AUC-IL', 'AUC-BLA', '% IL', '% BLA'])

#left2col = df.iloc[:,0]
AUC_IL = df.iloc[0,1]          # SHOULD BE THE IL DATA, FIRST ONE RECORDED FOR THAT RAT
AUC_BLA = df.iloc[0,2]         # SAME BUT FOR BLA DATA, EVERYTHING BELOW DEPENDS ON THAT KEY FACT

rows = len(df.iloc[:,1])

for x in range(rows):
    a = df.iloc[x,1]/AUC_IL
    ilPercent.append(a)
    b = df.iloc[x, 2]/AUC_BLA
    blaPercent.append(b)

df['% IL'] = ilPercent
df['% BLA'] = blaPercent

saveL = files_path + '\Percentages_Calculated_w_Graphs_' + RatSub + '.xlsx'
df.to_excel(saveL) #could probs use files_path

#UNDERNEATH GRAPH SAVING, SUCCESS FOR IL n BLA!
index = np.arange(rows)
bar_width = .9

wb = openpyxl.load_workbook(saveL)
ws = wb.active

#########################################################################################################IL
barIL = plt.bar(index, ilPercent, bar_width, align='edge', color = '#FFA500') #hex color orange
plt.xticks(index + bar_width / 2, ('PRE', '5', '30', '24','PRE','5','30','24','PRE','5','30','24'))
plt.title('ERP-IL Percentages')
plt.xlabel('Trials')
plt.ylabel('Percentage Change from Trial 1')

plt.savefig(files_path + "\IL_Percentage_" + RatSub + ".png")#, dpi = 150)

img1 = openpyxl.drawing.image.Image(files_path + "\IL_Percentage_" + RatSub + ".png")
img1.anchor = 'I1' #(ws.cell(column='I',row='2'))
ws.add_image(img1)
#close old plot so new doesn't overcolor
plt.close()
#########################################################################################################BLA
barBLA = plt.bar(index, blaPercent, bar_width, align='edge', color = '#000080') #hex navy blue
plt.xticks(index + bar_width / 2, ('PRE', '5', '30', '24','PRE','5','30','24','PRE','5','30','24'))
plt.title('ERP-BLA Percentages')
plt.xlabel('Trials')
plt.ylabel('Percentage Change from Trial 1')

plt.savefig(files_path + "\BLA_Percentage_" + RatSub + ".png")

img2 = openpyxl.drawing.image.Image(files_path + "\BLA_Percentage_" + RatSub + ".png")
img2.anchor = 'I27' #(ws.cell(column='I',row='2'))
ws.add_image(img2)

wb.save(saveL)

print("Yay Success! 4 files, \n 1 CSV with calculated data, \n 2 Graphs in png, 1 of IL and 1 of BLA measurements, \n and 1 EXCEL Data sheet with percentages calculated and graphs")
print("should be in " + files_path)